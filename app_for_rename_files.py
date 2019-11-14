import os
import re
import sys
from openpyxl import load_workbook
from PyQt5.QtWidgets import QWidget, QPushButton, QApplication, QLabel, QLineEdit, QMainWindow

lines = ('excel', 'coord_cells', 'files', 'separator', 'additional', 'file_extension')
lines_with_label = {lines[n]: lines[n]+'_label' for n in range(len(lines))}
main_text_line = ('Введите путь до файла Excel, затем до папки, где находятся '
                     'файлы для переименования\nв соответствии с элементами таблицы.'
                     'Файлы должны быть скачены/загружены в том\nпорядке, '
                     'в котором они и делались.')
forbidden_symbols = ("\\", "/", ":", "*", "?", '"', "<", ">", "|", "+", "%", "!", "@")

texts_line = ("Введите либо полный путь до файла Excel\n" \
                "либо просто имя файла:",
                "Введите сначала номер строки,\nс которой начинаются " \
                "нужные элементы\nтаблицы и после `;` буквы нужных колонок",
                "Путь к папке с файлами, " \
                "(в папке\nдолжны быть только нужные файлы)",
                "Символ(ы), разделяющий(-ие)\nсодержание колонок",
                "Дополнительное слово, которое будет\nв конце имени каждого файла, " \
                "например,\nчек", "Введите расширение файлов, которых\nхотите переименовать")
error_path_exc = "При вводе пути Excel возникла ошибка - вбейте другой путь до файла"
error_path_f = "При вводе директории возникла ошибка - вбейте другой путь до директории\nили проверьте расширение файлов"
error_extension = "Неверный формат расширения файла - проверьте"

error_sep = "Неверный формат разделителя - проверьте"

class Widget(QWidget):
    """	Main class for widget
	Make structure of programm"""
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setGeometry(350, 200, 510, 500)
        self.main_label = QLabel(self)
        self.main_label.move(20, 30)
        self.main_label.setText(main_text_line)
        self.coord_lines_y = 40

        for n, name_line in enumerate(lines):
            self.create_line(name_line)
            self.edit_args_line(name_main = lines[n],
                                name_label = lines_with_label[name_line],
                                text=texts_line[n])

        self.run_butt = QPushButton('Пуск!', self)
        self.run_butt.setGeometry(220, self.coord_lines_y+50, 60, 30)
	# Create line for input some args or path
    def create_line(self, name):
        setattr(self, name+'_label', QLabel(self))
        setattr(self, name, QLineEdit(self))
	# Edit view of elements of programm
    def edit_args_line(self, name_main, name_label, text, x_size = 240, y_size = 20):
        self.coord_lines_y += 60
        getattr(self, name_main).setGeometry(250, self.coord_lines_y, x_size, y_size)
        getattr(self, name_label).setText(text)
        getattr(self, name_label).move(20, self.coord_lines_y)


class Rename_files:
    """Rename all files at the directory by elements in Excel table
	You can match needed columns - there will use in new name of file."""
    def __init__(self):
        self.ui = Widget()
        self.ui.run_butt.clicked.connect(self.start_rename)

    def rename(self, ws):
        num_cell = int(self.ui.coord_cells.displayText().split(';')[0])
        letters_cell = [letter.upper() for letter in getattr(self.ui, "coord_cells").displayText().split(';')[1]]
        for f in self.sort_date_list:
            num_cell += 1
            separator = getattr(self.ui, "separator").displayText()
            text = separator.join([ws[i+str(num_cell)].value.replace('/', 'дробь') for i in letters_cell])
            text = text + separator +  getattr(self.ui, 'additional').displayText()
            if len(getattr(self.ui, 'file_extension').displayText()) < 3:
                getattr(self.ui, 'file_extension').setText(error_extension)
            else:
                os.rename(os.path.join(self.path_f, f[0]),
                          os.path.join(self.path_f, text + '.' +
                                       getattr(self.ui, 'file_extension').displayText().lower()))

    def start_rename(self):
        self.path_exc = getattr(self.ui, "excel").displayText().replace('\\', '/')
        if self.path_exc[-5:] != '.xlsx':
            self.path_exc+='.xlsx'
        has_forbidden_symbols = [i for i in getattr(self.ui, "separator").displayText() if i in forbidden_symbols]
        if len(getattr(self.ui, "separator").displayText()) > 4 or has_forbidden_symbols:
            getattr(self.ui, "separator").setText(error_sep)
            return
        self.path_f = getattr(self.ui, "files").displayText().replace('\\', '/')
        has_error_exc = 1
        for root, dirs, files in os.walk(self.path_exc[:self.path_exc.rfind('/')]):
            for file in files:
                if file == self.path_exc[self.path_exc.rfind('/')+1:]:
                    has_error_exc = 0
                    wb = load_workbook(filename=self.path_exc)
                    ws = wb.worksheets[0]
                    dir_list = [os.path.join(self.path_f, x) for x in os.listdir(self.path_f)]
                    has_error_f_list = [file_name for file_name in dir_list \
                                         if getattr(self.ui, 'file_extension').displayText().lower() in file_name]
                    if has_error_f_list:
                        date_list = [[x, os.path.getmtime(x)] for x in dir_list]
                        self.sort_date_list = sorted(date_list, key=lambda x: x[1])
                        self.rename(ws)
                    else:
                        getattr(self.ui, "files").setText(error_path_f)
                    return
        if has_error_exc:
            getattr(self.ui, "excel").setText(error_path_exc)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    myapp = Rename_files()
    myapp.ui.show()
    sys.exit(app.exec_())